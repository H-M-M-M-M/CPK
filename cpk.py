import streamlit as st
import pandas as pd
import os
from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook
from pandas import ExcelWriter
import barcode
from barcode.writer import ImageWriter
from PIL import Image, ImageDraw, ImageFont
from barcode import get_barcode_class


# 保存路径配置
SAVE_FOLDER = r"\\161.92.72.15\Test_Engineering\Files_Transfer\Joicy"
EXCEL_FILE = os.path.join(SAVE_FOLDER, "调胶记录.xlsx")
CACHE_FILE = os.path.join(SAVE_FOLDER, "buffer.csv")
SHEET_NAME = "Sheet1"

# 创建保存文件夹
os.makedirs(SAVE_FOLDER, exist_ok=True)

# 安全写入 Excel，支持文件占用时缓存
def append_to_excel_safely(excel_path, new_df, sheet_name="Sheet1", cache_file="buffer.csv"):
    if os.path.exists(cache_file):
        cache_df = pd.read_csv(cache_file)
        new_df = pd.concat([cache_df, new_df], ignore_index=True)
        os.remove(cache_file)

    expected_columns = [
        "记录ID", "操作员", "批次号", "胶水A型号", "胶水B型号", "胶水A重量", "胶水B重量",
        "有效期", "调胶时间", "备注"
    ]
    new_df = new_df[expected_columns]

    try:
        if not os.path.exists(excel_path):
            new_df.to_excel(excel_path, index=False, sheet_name=sheet_name)
        else:
            book = load_workbook(excel_path)
            with ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                start_row = writer.sheets[sheet_name].max_row
                new_df.to_excel(writer, startrow=start_row, index=False, header=False, sheet_name=sheet_name)
        st.success(f"✅ 提交成功！记录 ID 为：{new_df.iloc[-1]['记录ID']}")
    except PermissionError:
        new_df.to_csv(cache_file, index=False)
        st.warning("⚠️ Excel 文件被占用，已将记录缓存。请稍后关闭文件后刷新页面提交。")

# 页面标题
st.title("🧪 调胶记录提交系统")

# 表单输入
with st.form("glue_form"):
    operator = st.text_input("👨‍🔧 操作员")
    batch_number = st.text_input("🔢 批次号")
    glue_a_model = st.text_input("🧪 胶水 A 型号")
    glue_b_model = st.text_input("🧪 胶水 B 型号")
    glue_a_weight = st.number_input("⚖️ 胶水 A 重量 (g)", min_value=0.0, step=0.1)
    glue_b_weight = st.number_input("⚖️ 胶水 B 重量 (g)", min_value=0.0, step=0.1)
    valid_days = st.number_input("📅 有效期（天）", min_value=1, step=1, value= None)
    remark = st.text_area("📝 备注")
    submitted = st.form_submit_button("✅ 提交记录")

# 表单提交后操作
if submitted:
    now = datetime.now()
    record_id = now.strftime("%Y%m%d%H%M%S")
    record_time = now.strftime("%Y-%m-%d %H:%M:%S")
    #valid_date = now.date() + pd.to_timedelta(valid_days, unit="d")
    valid_date = now + pd.to_timedelta(valid_days, unit="d")


    # 写入 DataFrame
    new_df = pd.DataFrame([{
        "记录ID": record_id,
        "操作员": operator,
        "批次号": batch_number,
        "胶水A型号": glue_a_model,
        "胶水B型号": glue_b_model,
        "胶水A重量": glue_a_weight,
        "胶水B重量": glue_b_weight,
        "有效期": valid_date,
        "调胶时间": record_time,
        "备注": remark,
    }])

    append_to_excel_safely(EXCEL_FILE, new_df, SHEET_NAME, CACHE_FILE)

    # 生成高清条形码
    barcode_class = barcode.get_barcode_class('code128')
    barcode_obj = barcode_class(record_id, writer=ImageWriter())

    barcode_buffer = BytesIO()
    barcode_obj.write(barcode_buffer, {
        "module_width": 0.2,     # 紧凑线宽
        "module_height": 3.0,   # 高度适中
        "font_size": 7,          # 条码编号字体大小
        "text_distance": 3,    # 数字紧贴条形码
        "dpi": 300               # 高分辨率
    })
    barcode_buffer.seek(0)
    barcode_img = Image.open(barcode_buffer)

    # 准备下方文字
    text_lines = [
        f"调胶日期：{now.strftime('%Y-%m-%d %H:%M:%S')}",
        f"有效日期：{valid_date.strftime('%Y-%m-%d %H:%M:%S')}"
    ]
    font_path = "C:/Windows/Fonts/msyh.ttc"
    font = ImageFont.truetype(font_path, 24)

    # 计算文字区域尺寸
    line_heights = [font.getbbox(line)[3] for line in text_lines]
    text_height = sum(line_heights)
    text_width = max(font.getbbox(line)[2] for line in text_lines)

    # 合成新图：条形码在上，文字在下
    new_width = max(barcode_img.width, text_width) + 10
    new_height = barcode_img.height + text_height + 4  # 更紧凑
    new_img = Image.new("RGB", (new_width, new_height), "white")
    new_img.paste(barcode_img, ((new_width - barcode_img.width) // 2, 0))

    # 粘贴条形码
    new_img.paste(barcode_img, ((new_width - barcode_img.width) // 2, 0))

    # 写下方文字
    draw = ImageDraw.Draw(new_img)
    text_y = barcode_img.height
    for line in text_lines:
        text_x = (new_width - font.getbbox(line)[2]) // 2
        draw.text((text_x, text_y), line, font=font, fill="black")
        text_y += font.getbbox(line)[3] + 1  # 控制行距

    # 显示并下载条形码
    img_buffer = BytesIO()
    new_img.save(img_buffer, format="PNG")
    img_buffer.seek(0)

    st.image(img_buffer, caption=f"📎 记录 ID 条形码：{record_id}", use_column_width=False)
    st.download_button(
        label="📥 下载条形码图片（紧凑布局）",
        data=img_buffer,
        file_name=f"{record_id}_barcode_compact.png",
        mime="image/png"
    )
